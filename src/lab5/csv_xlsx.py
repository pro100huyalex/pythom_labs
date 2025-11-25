import csv
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX.
    """
    # Создаем Excel книгу и лист
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    
    with open(csv_path, 'r', encoding='utf-8') as file:
        rows = list(csv.reader(file))

    for row_num, row_data in enumerate(rows, 1):
        for col_num, value in enumerate(row_data, 1):
            sheet.cell(row=row_num, column=col_num, value=value)
    
    # Заголовки жирным
    for col in range(1, len(rows[0]) + 1):
        sheet.cell(row=1, column=col).font = Font(bold=True)
    
    for col in range(1, len(rows[0]) + 1):
        col_letter = get_column_letter(col)
        max_len = max((len(str(cell.value)) for cell in sheet[col_letter] if cell.value), default=0)
        sheet.column_dimensions[col_letter].width = max(8, max_len + 2)
    
    # Сохраняем и закрываем
    workbook.save(xlsx_path)
    workbook.close()


if __name__ == "__main__":
    # Тест
    test_data = [
        ["Имя", "Возраст", "Город"],
        ["Анна", "25", "Москва"],
        ["Иван", "30", "СПб"]
    ]
    
    with open('test.csv', 'w', encoding='utf-8', newline='') as f:
        csv.writer(f).writerows(test_data)
    
    csv_to_xlsx('test.csv', 'test.xlsx')
    print( "test.xlsx создан")